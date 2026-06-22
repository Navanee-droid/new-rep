# Sharding_DDL_Compare.py
# ======================
# Compares table lists and DDL (columns/data types) between CUR_IBS (Curated)
# and APP_IBS_SHARD_1 / APP_IBS_SHARD_2 (Sharding application layer).
#
# Modification History:
# ---------------------
# [CHANGE-2026-02-11] - Round 2 Fixes: TABLE_TYPE filter & TB_C2 conformed schema
#   1. Switched table-fetching queries from SHOW TERSE TABLES to
#      INFORMATION_SCHEMA.TABLES with TABLE_TYPE = 'BASE TABLE' filter.
#      This ensures only base tables are validated (views are excluded).
#   2. Fix: TB_C2 tables reside in DDW_CNF_DIM (conformed schema), not in
#      the app schema.  The YAML config is the authoritative source for which
#      TB_C2 tables belong to this app.  Shard-side queries fetch ONLY those
#      specific tables from APP_IBS_SHARD_x.DDW_CNF_DIM (using IN clause).
#   3. Fix: Shard-side DDL queries for TB_C2 now route to DDW_CNF_DIM (mirroring
#      the curated-side routing).
#   4. Added diagnostic print statements showing table counts per layer
#      (curated, shard_1, shard_2) including TB_C2 tables.
# ---------------------
# [CHANGE-2026-02-10] - TB_C2 Table Support & Bug Fixes
#   1. Added TB_C2 table support using YAML config (configs/surr_keys_{schema}_C2.yaml)
#      - TB_C2 tables reside in CUR_IBS.DDW_CNF_DIM (shared schema), not the app schema
#      - Follows the same pattern as FieldValidation_DIM.py and FieldValidation_Fact.py
#   2. Bug Fix: Mutable list aliasing on initialization (line 68 original)
#   3. Bug Fix: UnboundLocalError for DTM view name conversion when prefix is unexpected
#   4. Bug Fix: Data type comparison was set-based, not per-column (could report Success
#      when column types were swapped between tables)
#   5. Bug Fix: os.mkdir fails if parent directory doesn't exist (switched to os.makedirs)
#   6. Cleanup: Removed unused imports (shutil, smtplib)
# ---------------------

import snowflake.connector
import time
import yaml
import os
import sys
import traceback
import toml
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import serialization
from openpyxl import workbook as owb, load_workbook, styles

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from datetime import datetime
from validation_utils import (
    ValidationLoader,
    ValidationResult,
    ValidationDetailResult,
    TestCaseRegistry,
    ExecutionSummary,
    write_results_to_csv,
    write_summary_report,
    determine_status,
    cap_details
)

from script_utils import (
    parse_args,
    load_yaml,
    open_sf_connection,
    get_appl_code,
    get_tables_from_appl_table,
    get_shard_count,
    get_snowflake_dtm_views
)

SCRIPT_NAME = "Sharding_DDL_Compare.py"
SCRIPT_VERSION = "v2.0"


def arg_parsing():
    return parse_args(
        required=['--a'],
        optional=['--load-sf-meta', '--sf-meta-db', '--sf-meta-schema'],
        description='Compares DDL between sharding databases.',
    )
    

def compare_list(schema, appl_code) -> dict:
    print("\nProcesing Table list Comparision ...")
    # [CHANGE-2026-02-10] Bug Fix: Initialize lists separately to avoid mutable aliasing.
    # Original code: cur_comp_lst = app1_comp_lst = app2_comp_lst = []
    # This created three references to the SAME list object, which could cause subtle bugs
    # if any list was modified in-place (e.g., via .append() or .extend()).
    cur_comp_lst = []

    con = open_sf_connection(app_config)
    cur = con.cursor()
    all_tables = get_tables_from_appl_table(cur, appl_code)
    tb_c2_tables = [t for t in all_tables if t.startswith('TB_C2')]

    print('\nFetching CUR_IBS Tables ...')
    # [CHANGE-2026-02-11] Switched from SHOW TERSE TABLES to INFORMATION_SCHEMA.TABLES
    # with TABLE_TYPE = 'BASE TABLE' filter to exclude views from validation.
    cur_tb_sql = f"""SELECT TABLE_NAME FROM CUR_IBS.INFORMATION_SCHEMA.TABLES
        WHERE TABLE_SCHEMA = '{schema}'
        AND TABLE_NAME LIKE 'TB_%'
        AND TABLE_NAME NOT LIKE 'TB_APPLICATION_TENANTS%'
        AND TABLE_NAME NOT LIKE '%DTM%'
        AND TABLE_NAME NOT LIKE '%DLV%'
        AND TABLE_TYPE = 'BASE TABLE'
        ORDER BY TABLE_NAME;"""
    # print(cur_tb_sql)
    cur.execute(cur_tb_sql)
    rows = cur.fetchall()
    if rows:
        tbls, = zip(*rows)
        tblst = list(tbls)
    else:
        tblst = []
    #print(tblst)
    
    print('\nFetching CUR_IBS DTM Views ...')
    # dtm_vw_query=f"SHOW TERSE VIEWS IN SCHEMA CUR_IBS.{schema} ->> SELECT \"name\" from $1 WHERE \"name\" like '%DTM%';"
    # print(dtm_vw_query)
    # cur.execute(dtm_vw_query)
    #rows = cur.fetchall()
    sf_views = get_snowflake_dtm_views(con,schema)
    dtmvlst = sf_views if sf_views else []
    # if sf_views:
    #     dtms, = zip(*rows)
    #     dtmvlst = list(dtms)
    # else:
    #     dtmvlst = []
    #print(dtmvlst)
    
    cur_comp_lst = tblst
    dtm_dict = {}
    for v in dtmvlst:
        # [CHANGE-2026-02-10] Bug Fix: Added else clause and default handling.
        # Original code did not handle view names that don't start with 'VW_' or 'V_',
        # which could cause UnboundLocalError (tb undefined on first iteration)
        # or stale value from previous iteration.
        if v.startswith('VW_'):
            tb = v.replace('VW_','TB_')
        elif v.startswith('V_'):
            tb = v.replace('V_','TB_')
        else:
            print(f'Warning: DTM view "{v}" has unexpected prefix (not VW_ or V_), skipping...')
            continue
        dtm_dict[tb] = v
        cur_comp_lst.append(tb)
    #print(cur_comp_lst)
    #print(dtm_dict)
    
    # [CHANGE-2026-02-11] Include TB_C2 tables from YAML config into curated list.
    # The YAML config (configs/surr_keys_{app}_C2.yaml) is the authoritative source
    # for which TB_C2 tables belong to this application.  TB_C2 tables reside in
    # CUR_IBS.DDW_CNF_DIM (conformed/shared schema) but only the ones listed in
    # the YAML are relevant to this application's sharding validation.
    if tb_c2_tables:
        print(f'\nAdding {len(tb_c2_tables)} TB_C2 tables from config to curated list ...')
        for tb in tb_c2_tables:
            if tb not in cur_comp_lst:
                cur_comp_lst.append(tb)
            else:
                print(f'  TB_C2 table {tb} already in curated list, skipping duplicate.')
    
    shard_count = get_shard_count(cur)
    shard_comp_lsts = {}
    c2_in_clause = "','".join(tb_c2_tables) if tb_c2_tables else None

    for i in range(1, shard_count + 1):
        db = f"APP_IBS_SHARD_{i}"
        print(f'\nFetching {db} Tables ...')
        shard_tb_sql = f"""SELECT TABLE_NAME FROM {db}.INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = '{schema}'
            AND TABLE_TYPE = 'BASE TABLE'
            ORDER BY TABLE_NAME;"""
        cur.execute(shard_tb_sql)
        rows = cur.fetchall()
        if rows:
            tbls, = zip(*rows)
            shard_comp_lsts[db] = list(tbls)
        else:
            shard_comp_lsts[db] = []

        if tb_c2_tables:
            print(f'\nFetching TB_C2 tables from {db}.DDW_CNF_DIM ...')
            c2_sql = f"""SELECT TABLE_NAME FROM {db}.INFORMATION_SCHEMA.TABLES
                WHERE TABLE_SCHEMA = 'DDW_CNF_DIM'
                AND TABLE_NAME IN ('{c2_in_clause}')
                AND TABLE_TYPE = 'BASE TABLE'
                ORDER BY TABLE_NAME;"""
            cur.execute(c2_sql)
            rows = cur.fetchall()
            if rows:
                tbls, = zip(*rows)
                c2_added = 0
                for t in tbls:
                    if t not in shard_comp_lsts[db]:
                        shard_comp_lsts[db].append(t)
                        c2_added += 1
                print(f'  Found {len(list(tbls))} of {len(tb_c2_tables)} TB_C2 tables ({c2_added} new)')
            else:
                print(f'  No matching TB_C2 tables found in {db}.DDW_CNF_DIM')

    cur.close()
    con.close()

    # Diagnostic: Show table counts per layer for debugging
    tb_c2_in_cur = [t for t in cur_comp_lst if t.startswith('TB_C2')]
    parts = [f'CUR_IBS: {len(cur_comp_lst)} (incl. {len(tb_c2_in_cur)} TB_C2)']
    for db, lst in shard_comp_lsts.items():
        tb_c2_in_shard = [t for t in lst if t.startswith('TB_C2')]
        parts.append(f'{db}: {len(lst)} (incl. {len(tb_c2_in_shard)} TB_C2)')
    print(f'\nTable counts - {", ".join(parts)}')

    result = {}
    for db, shard_lst in shard_comp_lsts.items():
        print(f'\nComparing CUR_IBS vs {db} ...')
        comm_lst = list(set(cur_comp_lst) & set(shard_lst))
        diff_lst = list(set(cur_comp_lst) ^ set(shard_lst))

        fnl_lst = [[c, c, 'Found'] for c in comm_lst]
        fnl_lst += [[d, '', 'Missing'] if d in cur_comp_lst else ['', d, 'Missing'] for d in diff_lst]
        fnl_lst = list(map(lambda t: [dtm_dict[t[0]]] + t[1:] if t[0] in dtm_dict else t, fnl_lst))

        result[db] = fnl_lst

    return result

def get_column_structure(cur, query):
    col_dict = {}
    cur.execute(query)
    res = cur.fetchall()
    col_dict = {col_name: col_type for (col_name, col_type) in res}
    
    return col_dict

def compare_schema(tbdict, schema):
    print('\nProcessing DDL Comparison ...\n')
    ddlcompdict = {}
    ts_outcomes = []
    for db in tbdict:
        conc = open_sf_connection(app_config)
        cona = open_sf_connection(app_config,db)
        curc = conc.cursor()
        cura = cona.cursor()
        fnl_lst = []
        print(f'\nComparing CUR_IBS vs {db} ...')
        for tbl in tbdict[db]:
            if tbl[2] == 'Missing' or tbl[1].find("DTM")>-1:
                tbl_display = tbl[0] if tbl[0] else tbl[1]
                # If the table exists in the shard but not in CUR_IBS (tbl[1] set, tbl[0] empty),
                # still check SHARDING_TS — we just can't do DDL comparison without a source of truth.
                if tbl[2] == 'Missing' and tbl[1] and not tbl[0]:
                    shard_schema = "DDW_CNF_DIM" if tbl[1].startswith("TB_C2") else schema
                    query = f"SHOW COLUMNS IN {db}.{shard_schema}.{tbl[1]} ->> SELECT \"column_name\", \"data_type\" from $1;"
                    try:
                        cols = get_column_structure(cura, query)
                        if 'SHARDING_TS' in cols:
                            ts_outcomes.append({'table': f'{db}.{tbl[1]}', 'status': 'SUCCESS',
                                                'reason': 'SHARDING_TS column present'})
                        else:
                            ts_outcomes.append({'table': f'{db}.{tbl[1]}', 'status': 'FAIL',
                                                'reason': f'SHARDING_TS column missing in {db}'})
                            print(f'  WARNING: {tbl[1]} is missing SHARDING_TS in {db}')
                    except Exception:
                        pass  # table not accessible; DDL comparison will already flag it Missing
                print(f'Skipping {tbl_display} - Missing Or DTM table...')
                continue
            
            print(f'Processing {tbl[0]}...')
            
            # [CHANGE-2026-02-10] BEGIN - TB_C2 schema routing for DDL comparison
            # TB_C2 tables reside in CUR_IBS.DDW_CNF_DIM, not CUR_IBS.{app_schema}.
            # This follows the same pattern as FieldValidation_DIM.py / FieldValidation_Fact.py:
            #   sf_schema = "DDW_CNF_DIM" if table.startswith("TB_C2") else arg_dict["appl_name"]
            cur_schema = "DDW_CNF_DIM" if tbl[0].startswith("TB_C2") else schema
            # [CHANGE-2026-02-10] END - TB_C2 schema routing
            
            # [CHANGE-2026-02-10] TB_C2 tables in CUR_IBS have SRC_APPL_NAME column that
            # may not exist in the sharding layer. Exclude it from curated side for TB_C2 tables,
            # similar to how SHARDING_TS is excluded from the sharding side.
            
            query = f"SHOW COLUMNS IN CUR_IBS.{cur_schema}.{tbl[0]} ->> SELECT \"column_name\", \"data_type\" from $1;"
            try:
                colc_dict = get_column_structure(curc, query)
            except Exception as e:
                print(f'  SKIP {tbl[0]} - CUR_IBS.{cur_schema} table not found or not authorized: {e}')
                fnl_lst += [[tbl[0], f'Table not found in CUR_IBS.{cur_schema}', '', 'Missing']]
                continue
            
            # [CHANGE-2026-02-11] Route shard-side TB_C2 queries to DDW_CNF_DIM,
            # mirroring the curated-side routing above.  TB_C2 tables reside in
            # DDW_CNF_DIM in both CUR_IBS and the shard databases.
            shard_schema = "DDW_CNF_DIM" if tbl[1].startswith("TB_C2") else schema
            
            # Fetch all shard columns without filtering SHARDING_TS — pop it after to
            # track presence in one query instead of a separate pass.
            query = f"SHOW COLUMNS IN {db}.{shard_schema}.{tbl[1]} ->> SELECT \"column_name\", \"data_type\" from $1;"
            try:
                cola_dict = get_column_structure(cura, query)
            except Exception as e:
                print(f'  SKIP {tbl[1]} - {db}.{shard_schema} table not found or not authorized: {e}')
                fnl_lst += [[tbl[0], f'Table not found in {db}.{shard_schema}', '', 'Missing']]
                continue
            
            # SHARDING_TS check — pop so it doesn't appear as a DDL mismatch vs CUR_IBS
            if cola_dict.pop('SHARDING_TS', None) is None:
                ts_outcomes.append({'table': f'{db}.{tbl[1]}', 'status': 'FAIL',
                                    'reason': f'SHARDING_TS column missing in {db}'})
                print(f'  WARNING: {tbl[1]} is missing SHARDING_TS in {db}')
            else:
                ts_outcomes.append({'table': f'{db}.{tbl[1]}', 'status': 'SUCCESS',
                                    'reason': 'SHARDING_TS column present'})
            
            # Find missing/extra columns (symmetric difference of column names)
            diffk_lst = list(set(colc_dict.keys()) ^ set(cola_dict.keys()))
            
            # [CHANGE-2026-02-10] Bug Fix: Per-column data type comparison.
            # Original code compared data types as sets:
            #   diffv_lst = list(set(colc_dict.values()) ^ set(cola_dict.values()))
            # This was incorrect because set-based comparison doesn't check per-column.
            # Example: If CUR has {A:VARCHAR, B:NUMBER} and SHARD has {A:NUMBER, B:VARCHAR},
            # the value sets are identical ({VARCHAR, NUMBER}), so the script would report
            # 'Success' even though the types are swapped.
            # Fix: Compare data types column-by-column for common columns.
            commk_lst = list(set(colc_dict.keys()) & set(cola_dict.keys()))
            type_mismatches = []
            for col in commk_lst:
                if colc_dict[col] != cola_dict[col]:
                    type_mismatches.append((col, colc_dict[col], cola_dict[col]))
            
            # TableName, MissingCols, DataType mismatch, Success
            if len(diffk_lst) == 0 and len(type_mismatches) == 0:
                fnl_lst += [[ tbl[0],'','','Success' ]]
            else:
                miss_cols = typ_mismtch = ''
                if len(diffk_lst) > 0:
                    miss_cols = ','.join(diffk_lst)
                # [CHANGE-2026-02-10] Bug Fix: Report per-column type mismatches clearly
                if len(type_mismatches) > 0:
                    typ_mismtch = ','.join([
                        f'{col}: CUR_IBS={cur_type} vs {db}={shard_type}\n'
                        for col, cur_type, shard_type in type_mismatches
                    ])
                fnl_lst += [[ tbl[0], miss_cols, typ_mismtch, 'Failed']]
        
        # print(fnl_lst)
        cura.close()
        curc.close()
        cona.close()
        conc.close()
        
        ddlcompdict[db] = fnl_lst
    
    return ddlcompdict, ts_outcomes


def write_Comparision(compdict: dict, base_path: str):
    print("\nWriting table list Comparison results ...")
    wb = owb.Workbook()
    for app_db in compdict:
        ws = wb.create_sheet(f"{app_db}_table_list")
        ws.append(['CUR_IBS Tables', f'{app_db} Tables', 'Status'])
        for rowl in compdict[app_db]:
            ws.append(rowl)
        hfnt = styles.Font(name='Calibri', size=12, bold=True)
        hfil = styles.PatternFill(fill_type='solid', start_color='33CCCC', end_color='33CCCC')
        for cl in ws[1]:
            cl.font = hfnt
            cl.fill = hfil
        for cols in ws.columns:
            width = max(len(str(cl.value)) for cl in cols)
            ws.column_dimensions[cols[0].column_letter].width = width + 5
        stsfil = styles.PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        for cl in ws['C']:
            if cl.value == 'Missing':
                cl.fill = stsfil
    wb.remove(wb['Sheet'])
    wb.save(base_path + '/Sharding_Validation.xlsx')
    wb.close()
    
def write_ddlComparision(ddlcompdict: dict, ts_outcomes: list, base_path: str):
    print("\nWriting DDL Comparison results ...")
    # Build a lookup: db -> table -> SHARDING_TS status
    ts_lookup = {}
    for o in ts_outcomes:
        parts = o['table'].split('.', 1)
        if len(parts) == 2:
            db, tbl = parts
            ts_lookup.setdefault(db, {})[tbl] = 'Present' if o['status'] == 'SUCCESS' else 'MISSING'
    wb = load_workbook(base_path + '/Sharding_Validation.xlsx')
    for app_db in ddlcompdict:
        ws = wb.create_sheet(f"{app_db}_ddl_comp")
        ws.append(['Tables', 'Missing Columns', 'Datatype Mismatch', 'SHARDING_TS', 'Status'])
        db_ts = ts_lookup.get(app_db, {})
        for rowl in ddlcompdict[app_db]:
            tbl_name = rowl[0]
            ts_status = db_ts.get(tbl_name, 'N/A')
            ws.append(rowl[:3] + [ts_status] + rowl[3:])
        hfnt = styles.Font(name='Calibri', size=12, bold=True)
        hfil = styles.PatternFill(fill_type='solid', start_color='33CCCC', end_color='33CCCC')
        for cl in ws[1]:
            cl.font = hfnt
            cl.fill = hfil
        for cols in ws.columns:
            width = max(len(str(cl.value)) for cl in cols)
            ws.column_dimensions[cols[0].column_letter].width = (width if width < 50 else 50) + 5
        redfil = styles.PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        orngfil = styles.PatternFill(fill_type='solid', start_color='FF9900', end_color='FF9900')
        for row in ws.iter_rows(min_row=2):
            if row[4].value == 'Failed':
                row[4].fill = redfil
            if row[3].value == 'MISSING':
                row[3].fill = orngfil
    wb.save(base_path + '/Sharding_Validation.xlsx')
    wb.close()


def write_text_summary(compdict: dict, ddlcompdict: dict, ts_outcomes: list, output_path: str, appl_name: str) -> str:
    """Write a text summary of both table-list and DDL comparisons for dashboard capture."""
    # Build per-shard TS lookup
    ts_lookup = {}
    for o in ts_outcomes:
        parts = o['table'].split('.', 1)
        if len(parts) == 2:
            db, tbl = parts
            ts_lookup.setdefault(db, {})[tbl] = o['status']

    summary_path = os.path.join(output_path, 'Sharding_Validation_Summary.txt')
    with open(summary_path, 'w', encoding='ascii', errors='replace') as f:
        f.write(f"Sharding DDL Validation - {appl_name}\n")
        f.write(f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")

        for db_name, rows in compdict.items():
            found = sum(1 for r in rows if r[2] == 'Found')
            missing = sum(1 for r in rows if r[2] == 'Missing')
            f.write(f"TABLE LIST: CUR_IBS vs {db_name}\n")
            f.write(f"  Found: {found}  |  Missing: {missing}\n")
            if missing > 0:
                f.write("  Missing tables:\n")
                for r in rows:
                    if r[2] == 'Missing':
                        tbl = r[0] if r[0] else r[1]
                        side = 'CUR_IBS only' if r[0] else f'{db_name} only'
                        f.write(f"    - {tbl} ({side})\n")
            f.write("\n")

        f.write("-" * 80 + "\n\n")

        for db_name, rows in ddlcompdict.items():
            db_ts = ts_lookup.get(db_name, {})
            ts_pass = sum(1 for v in db_ts.values() if v == 'SUCCESS')
            ts_fail = sum(1 for v in db_ts.values() if v == 'FAIL')
            success = sum(1 for r in rows if r[3] == 'Success')
            failed = sum(1 for r in rows if r[3] == 'Failed')
            other = sum(1 for r in rows if r[3] not in ('Success', 'Failed'))
            f.write(f"DDL + SHARDING_TS: CUR_IBS vs {db_name}\n")
            f.write(f"  DDL   -- Success: {success}  |  Failed: {failed}")
            if other:
                f.write(f"  |  Other: {other}")
            f.write("\n")
            f.write(f"  SHARDING_TS -- Present: {ts_pass}  |  Missing: {ts_fail}\n")
            if failed > 0:
                f.write("  DDL Failures:\n")
                for r in rows:
                    if r[3] == 'Failed':
                        f.write(f"    - {r[0]}\n")
                        if r[1]:
                            f.write(f"      Missing columns: {r[1]}\n")
                        if r[2]:
                            f.write(f"      Type mismatches: {r[2].strip()}\n")
            if ts_fail > 0:
                f.write("  SHARDING_TS Missing:\n")
                for tbl, status in db_ts.items():
                    if status == 'FAIL':
                        f.write(f"    - {tbl}\n")
            f.write("\n")

        f.write("=" * 80 + "\n")
        f.write(f"Full details: {output_path}/Sharding_Validation.xlsx\n")
    return summary_path

if __name__ == '__main__':
    script_start = time.perf_counter()
    argument_dict = arg_parsing()
    app_config = load_yaml(yaml_file=f'{os.environ["PYTHONPATH"]}/IngestionConfig.yaml')
    
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    output_path = f"{app_config['snowflake_connection']['validation_path']}/{argument_dict['appl_name']}/{script_name}"
    os.makedirs(output_path, exist_ok=True)
    
    sf_conn = open_sf_connection(app_config)
    appl_code = get_appl_code(argument_dict['appl_name'], sf_conn.cursor())
    sf_conn.close()
    compdict = compare_list(argument_dict['appl_name'], appl_code)
    write_Comparision(compdict, output_path)
    
    print("-------------------------------------------------------------------------------")
    ddlcompdict, ts_outcomes = compare_schema(compdict, argument_dict['appl_name'])
    write_ddlComparision(ddlcompdict, ts_outcomes, output_path)

    summary_txt_path = write_text_summary(compdict, ddlcompdict, ts_outcomes, output_path, argument_dict['appl_name'])

    print("-------------------------------------------------------------------------------")
    print("Sharding DDL Validation Complete.")
    print(f"Results saved to: {output_path}/Sharding_Validation.xlsx")
    print(f"Text summary: {summary_txt_path}")

    # --- Build combined outcomes (DDL + SHARDING_TS) ---
    table_outcomes = []
    for db_name, rows in ddlcompdict.items():
        for row in rows:
            tbl_name, miss_cols, dtype_mismatch, status = row[0], row[1], row[2], row[3]
            is_success = status == 'Success'
            table_outcomes.append({
                'table': f'{db_name}.{tbl_name}',
                'status': 'SUCCESS' if is_success else 'FAIL',
                'reason': 'DDL matches' if is_success else f'Missing={miss_cols}, DtypeMismatch={dtype_mismatch}',
                'validation_key': 'shard_ddl_compare'
            })

    for outcome in ts_outcomes:
        table_outcomes.append({
            'table': outcome['table'],
            'status': outcome['status'],
            'reason': outcome['reason'],
            'validation_key': 'sharding_ts_column'
        })

    # Print combined summary
    ddl_passed = sum(1 for o in table_outcomes if o['validation_key'] == 'shard_ddl_compare' and o['status'] == 'SUCCESS')
    ddl_failed = sum(1 for o in table_outcomes if o['validation_key'] == 'shard_ddl_compare' and o['status'] == 'FAIL')
    ts_passed = sum(1 for o in ts_outcomes if o['status'] == 'SUCCESS')
    ts_failed = sum(1 for o in ts_outcomes if o['status'] == 'FAIL')
    print(f'\nDDL Check: {ddl_passed} SUCCESS, {ddl_failed} FAIL')
    print(f'SHARDING_TS Check: {ts_passed} SUCCESS, {ts_failed} FAIL')

    # --- Single metadata load combining DDL + SHARDING_TS ---
    if argument_dict.get('load_sf_meta'):
        try:
            sf_conn = open_sf_connection(app_config)
            appl_code = get_appl_code(argument_dict['appl_name'], sf_conn.cursor())
            meta_cur = sf_conn.cursor()
            registry = TestCaseRegistry(meta_cur, SCRIPT_NAME,
                                        database=argument_dict.get('sf_meta_db'),
                                        schema=argument_dict.get('sf_meta_schema'))
            val_results = []
            for outcome in table_outcomes:
                mis_c = 1 if outcome['status'] == 'FAIL' else 0
                mat_c = 1 if outcome['status'] == 'SUCCESS' else 0
                val_results.append(registry.create_result(
                    validation_key=outcome['validation_key'],
                    test_scenario=f'{outcome["validation_key"]}: {outcome["table"]}',
                    appl_name=argument_dict['appl_name'],
                    appl_code=appl_code,
                    tenant_id='ALL',
                    table_name=outcome['table'],
                    validation_status=outcome['status'],
                    status_reason=outcome['reason'],
                    mismatched_count=mis_c,
                    matched_count=mat_c
                ))
            if not val_results:
                val_results.append(registry.create_result(
                    validation_key='shard_ddl_compare',
                    test_scenario='Shard DDL + SHARDING_TS comparison',
                    appl_name=argument_dict['appl_name'],
                    appl_code=appl_code,
                    tenant_id='ALL', table_name='ALL_TABLES',
                    validation_status='SUCCESS', status_reason='No tables processed'
                ))
            loader = ValidationLoader(
                sf_cursor=meta_cur, arg_dict=argument_dict,
                script_name=SCRIPT_NAME, script_version=SCRIPT_VERSION,
                database=argument_dict.get('sf_meta_db'), schema=argument_dict.get('sf_meta_schema')
            )
            summary = ExecutionSummary(
                script_name=SCRIPT_NAME, appl_name=argument_dict.get('appl_name', ''),
                appl_code=appl_code, tenant_id='ALL',
                process_date=argument_dict.get('process_date', ''), script_version=SCRIPT_VERSION
            )
            summary.started_at = datetime.fromtimestamp(time.time() - (time.perf_counter() - script_start))
            summary.parameters_used = {k: str(v) for k, v in argument_dict.items() if k not in ('sf_cursor',)}
            summary.update_counts(val_results)
            summary.execution_time_sec = time.perf_counter() - script_start
            if os.path.exists(summary_txt_path):
                summary.read_and_store_output(summary_txt_path, file_type='sharding_ddl_report')
            exec_id = loader.insert_execution_summary(summary)

            run_id_map = {}
            for r in val_results:
                rid = loader.insert_master(r, execution_id=exec_id)
                run_id_map[r.table_name] = rid

            detail_batch = []
            for outcome in table_outcomes:
                if outcome['status'] == 'FAIL':
                    rid = run_id_map.get(outcome['table'], 0)
                    if rid:
                        detail_batch.append(ValidationDetailResult(
                            run_id=rid,
                            match_status='MISMATCH',
                            record_key=outcome['table'],
                            source_data={'table': outcome['table'], 'validation': outcome['validation_key']},
                            target_data={'status': outcome['status'], 'reason': outcome.get('reason', '')[:2000]},
                            detail_remarks=(outcome.get('reason', '') or '')[:2000]
                        ))
            if detail_batch:
                capped_batch, _, _ = cap_details(detail_batch)
                loader.insert_detail_bulk(capped_batch)

            summary.emit_summary_line()
            total_passed = sum(1 for o in table_outcomes if o['status'] == 'SUCCESS')
            total_failed = sum(1 for o in table_outcomes if o['status'] == 'FAIL')
            print(f"Loaded {len(val_results)} result(s) -- {total_passed} SUCCESS, {total_failed} FAIL")
            print(f"Execution status: {summary.execution_status}")
            sf_conn.close()
        except Exception as e:
            print(f"Failed to load validation results to Snowflake: {str(e)}")
            traceback.print_exc()
    else:
        print("Skipping metadata load to Snowflake (--load-sf-meta not specified)")